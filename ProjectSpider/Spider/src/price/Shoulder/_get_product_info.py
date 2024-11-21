# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from bs4 import BeautifulSoup
import requests
import glob
import os

from time import sleep
import random

import pandas as pd
from openpyxl import load_workbook

# Lista de User-Agents
USER_AGENTS = [
  "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
  "Mozilla/5.0 (X11;U; Linux i686; en-GB; rv:1.9.1) Gecko/20090624 Ubuntu/9.04 (jaunty) Firefox/3.5",
  "Mozilla/5.0 (Windows; U; Windows NT 5.1; de-DE; rv:1.9.2.20) Gecko/20110803 Firefox",
  ]
user_agent = random.choice(USER_AGENTS)

options = webdriver.ChromeOptions()
options.add_argument(f"user-agent={user_agent}")

# Headless
# options.add_argument("--headless")

# For debug
options.add_argument('--ignore-certificate-errors')
options.add_argument('--ignore-ssl-errors')
options.add_argument("--incognito")
options.add_argument("--start-maximized")
options.add_experimental_option('excludeSwitches', ['enable-logging'])


driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 10) 

def get_info(driver, url, row):
    driver.get(url)
    sleep(5)

    if row == 2:
      try:
        _path = "vtex-flex-layout-0-x-flexRow--container-product"
        wait.until(EC.presence_of_element_located((By.XPATH, _path)))
        sleep(2)  
      except:
        pass

    try:
      # Nome
      try:
        _path = "//h1[@class='shoulder-shoulder-app-11-x-productNamePDP']"
        _name = driver.find_element(By.XPATH, _path).get_attribute('innerText')
      except:
        _name = ''

      # preco actual
      try:
        _path = "//span[contains(@class, 'shoulder-shoulder-app-11-x-pricesPDP')]"
        _preco_sale = driver.find_element(By.XPATH, _path).get_attribute('innerText')
      except:
        _preco_sale = ''

      # preco original
      try:
        _path = "//span[contains(@class, 'shoulder-shoulder-app-11-x-listPricePDP')]"
        _preco_original = driver.find_element(By.XPATH, _path).get_attribute('innerText')
      except:
        _preco_original = _preco_sale

      # imagem
      try:
        _path = "//img[contains(@class, 'images-pdp')]"
        _img = driver.find_element(By.XPATH, _path).get_attribute('src')
      except:
        _img = ''

      # descricao
      try:
        _path = "//div[contains(@class, 'vtex-store-components-3-x-productDescriptionText')]"
        _desc = driver.find_element(By.XPATH, _path).get_attribute('innerText')
      except:
        _desc = ''

      # breadcrumb
      try:
        _path = "//div[contains(@data-testid,'breadcrumb')]//span//a"
        _breadcrumb = driver.find_elements(By.XPATH, _path)
        _categoria = [x.get_attribute('innerText') for x in _breadcrumb]
      except:
        _categoria = []

      print(_name, _preco_sale, _preco_original, _categoria ,_img, _desc)

      # driver.quit()

      return (_name, _preco_sale, _preco_original, ' '.join(_categoria) , _desc, _img)
    except:
      print('not found')
      # driver.quit()
      return ('', '', '', '', '')

if __name__ == "__main__":

  _marca = "Shoulder"

  _f = glob.glob(f"src//price//{_marca}//to_crawl_*.xlsx")
  files = [(f, os.path.basename(f).split('_')[3].split('.')[0]) for f in _f]
  print(files)

  # DEBUG
  # files = [
  #   (f'src/price/{_marca}/to_crawl_feminino_alfaiataria.xlsx', 'alfaiataria'),
  # ]

  for f,n in files:
    wb = load_workbook(f)
    ws = wb.active

    ws['E1'] = 'nome'
    ws['F1'] = 'preco_sale'
    ws['G1'] = 'preco_original'
    ws['H1'] = 'breadcrumb'
    ws['I1'] = 'descricao'
    ws['J1'] = 'img'

    row=2
    col=1
    while ws.cell(row, 3).value is not None:
      url = ws.cell(row, 3).value
      print(row-1,url)
      _nome, _preco_sale, _preco_original, _categoria, _desc, _img = get_info(driver, url, row)
      ws.cell(row,4).value = _nome
      ws.cell(row,5).value = _preco_sale
      ws.cell(row,6).value = _preco_original
      ws.cell(row,7).value = _categoria
      ws.cell(row,8).value = _desc
      ws.cell(row,9).value = _img

      row = row + 1

    wb.save(f"src/price/{_marca}/{n}.xlsx")
    driver.quit()