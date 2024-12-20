# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import json

from time import sleep
import random

import pandas as pd
from openpyxl import load_workbook

# Lista de User-Agents
USER_AGENTS = [
  "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
  "Mozilla/5.0 (X11;U; Linux i686; en-GB; rv:1.9.1) Gecko/20090624 Ubuntu/9.04 (jaunty) Firefox/3.5",
  "Mozilla/5.0 (Windows; U; Windows NT 5.1; de-DE; rv:1.9.2.20) Gecko/20110803 Firefox",
  ]
user_agent = random.choice(USER_AGENTS)

options = webdriver.ChromeOptions()
options.add_argument(f"user-agent={user_agent}")

# Headless
# options.add_argument("--headless")

# For debug
options.add_argument('--ignore-certificate-errors')
options.add_argument('--ignore-ssl-errors')
options.add_argument("--incognito")
options.add_argument("--start-maximized")
options.add_experimental_option('excludeSwitches', ['enable-logging'])


driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 10) 

def get_info(driver, url, row):
    driver.get(url)
    sleep(7)

    try:
      # Nome
      try:
        _path = "//span[contains(@class,'styles__ProductName-sc-ekm9ae-0')]"
        _name = driver.find_element(By.XPATH, _path).get_attribute('innerText')
      except:
        _name = ''

      # preco actual
      try:
        _path = "//span[@data-value = 'low-price']"
        _preco_sale = driver.find_element(By.XPATH, _path).get_attribute('innerText')
        _preco_sale = _preco_sale.split(' |')[0]
      except:
        _preco_sale = ''

      # preco original
      try:
        _path = "//span[@data-value = 'list-price']"
        _preco_original = driver.find_element(By.XPATH, _path).get_attribute('innerText')
      except:
        _preco_original = _preco_sale

      # imagem
      try:
        _path = "//div[contains(@class,'styles__ImageContainer-sc-1k59841-4')]//img"
        _img = driver.find_element(By.XPATH, _path).get_attribute('src')
      except:
        _img = ''

      # descricao
      try:
        _path = "//p[contains(@class, 'styles__ProductDescriptionText-sc-1xvn9i7-3')]"
        _desc = driver.find_element(By.XPATH, _path).get_attribute('innerText')
      except:
        try:
          _path = "//p[contains(@class, 'styles__ProductDescriptionText-sc-1xvn9i7-3')]"
          _desc = driver.find_element(By.XPATH, _path).get_attribute('innerText')
        except:
          _desc = ''

      # breadcrumb
      try:
        _path = "//div[contains(@class,'styles__BreadcrumbWrapper-sc-149j3f9-0')]//a"
        _breadcrumb = driver.find_elements(By.XPATH, _path)
        _breadcrumb = [x.get_attribute('innerText') for x in _breadcrumb]
      except:
        _breadcrumb = []

      print(_name, _preco_sale, _preco_original, _breadcrumb ,_img, _desc)

      return (_name, _preco_sale, _preco_original, ' '.join(_breadcrumb) , _desc, _img, 1)
    except:
      print('not found')
      return ('', '', '', '', '', '', 0)

if __name__ == "__main__":

  marca = "MariaFilo"

  # Arquivo com as informações para o crawler
  file = f'src/price/{marca}/to_crawl.xlsx'

  # Carrega os dados do excel
  wb = load_workbook(file)
  ws = wb.active

  row=2
  col=1
  while ws.cell(row, 3).value is not None:
    
    # Dados da linha da planilha
    catalogo = ws.cell(row,4).value
    nome = ws.cell(row,1).value
    url = ws.cell(row, 2).value

    print(catalogo, nome, url)

    # se o campo flg_crawled = 0, então busca a informação
    if ws.cell(row,5).value==0:

      # pega a URL para criar um ID
      product_id = url.split('/')[-2]

      # Retorna as informações da URL
      _nome, _preco_final, _preco_original, _categoria, _desc, _img, _flg = get_info(driver, url, row)

      # Dicionario de dados
      product_info = {
        'marca': marca,
        'catalogo': catalogo,
        'nome': _nome,
        'url': url,
        'preco_original': ''.join(filter(lambda x: x.isprintable(), _preco_original)),
        'preco_final': ''.join(filter(lambda x: x.isprintable(), _preco_final)),
        'breadcrumb': ''.join(filter(lambda x: x.isprintable(), _categoria)),
        'descricao': ''.join(filter(lambda x: x.isprintable(), _desc)),
        'imagem': _img
      }

      # Convert and write JSON object to file
      with open(f"src/price/{marca}/products/{product_id}.json", "w", encoding='utf-8') as f: 
          json.dump(product_info, f, ensure_ascii=False, indent= 4)

      # Marca como 1 o campo flg_crawled
      # ws.cell(row,5).value=_flg
      wb.save(file)
    
    # Proxima linha
    row = row + 1

  # Fecha o navegador
  driver.quit()